# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.2'
#       jupytext_version: 1.1.1
#   kernelspec:
#     display_name: python3_custom_julia
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Imports

# %%
# %load_ext autoreload
# %autoreload 2

import numpy as np
import pandas as pd

from tqdm.autonotebook import tqdm

# %%
import logging

_log = logging.getLogger()
_log.setLevel(logging.INFO)

# %%
import oil.utils.helpers as helpers
helpers.init_logger(logging.INFO)

# %%
from pylab import rcParams
rcParams['figure.figsize'] = (20, 5)

# %%
import data.data_config as data_config

# %%
import infra.helpers.telegram_notify.telegram_notify as tg

tgn = tg.TelegramNotify()

# %% [markdown]
# # Extract hyperlinks

# %%
import openpyxl

# %%
wb = openpyxl.load_workbook('/data/prices/product_slate_export_20190904.xlsx')

# %%
ws = wb[wb.get_sheet_names()[0]]

# %%
ws.cell(row=2, column=5).hyperlink.target

# %%
hyperlinks = []
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
    hyperlinks.append(row[0].hyperlink.target)

# %% [markdown]
# # Read excel

# %%
pse = pd.read_excel('/data/prices/product_slate_export_20190904.xlsx')

# %%
pse.head()

# %%
pse['product_link'] = hyperlinks

# %%
pse.head()

# %%
pse.tail()

# %%
pse.shape

# %% [markdown]
# # Save csv with hyperlinks

# %%
pse.to_csv('/data/prices/product_slate_export_20190904.csv', index=0)

# %%
